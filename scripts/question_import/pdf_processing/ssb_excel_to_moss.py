"""
excel_to_moss.py

Converts the SSB Question Writing Excel spreadsheet into MOSS-compatible
JSON packet files, one per round.

Blue-highlighted rounds in the Compiling sheet are the ones we export.
Category and question type are derived from the formula references in the
Compiling sheet (e.g. =Bio!D2 → BIOLOGY, column D = TOSSUP).

Usage:
    python excel_to_moss.py
    python excel_to_moss.py --input "path/to/file.xlsx" --output ./output
"""

import argparse
import json
import os
import re
import sys

import openpyxl
from openpyxl.utils import column_index_from_string

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "SSB Question Writing 2026.xlsx")
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "output")

# Fill color of blue-highlighted (wanted) rounds in the Compiling sheet col A
BLUE_FILL = "FFCFE2F3"

# Maps sheet name in formula → MOSS category string
SHEET_TO_CATEGORY = {
    "Bio": "BIOLOGY",
    "Chem": "CHEMISTRY",
    "ESS": "EARTH_SPACE",
    "Math": "MATH",
    "Physics": "PHYSICS",
    "Energy": "ENERGY",
}


# Subject sheet column structure: D=Tossup1, E=Bonus1, F=Tossup2, G=Bonus2, ...
# Column D has openpyxl 1-indexed value 4.
# Even offset from D (0, 2, 4...) = Tossup; odd (1, 3, 5...) = Bonus
def col_to_question_type(col_letter: str) -> str:
    col_1idx = column_index_from_string(col_letter)
    d_1idx = column_index_from_string("D")  # = 4
    return "TOSSUP" if (col_1idx - d_1idx) % 2 == 0 else "BONUS"


def smartify_quotes(text: str) -> str:
    """Convert straight ASCII quotes to directional Unicode quotes."""
    result = []
    in_double = False
    for i, ch in enumerate(text):
        if ch == '"':
            result.append('\u201d' if in_double else '\u201c')
            in_double = not in_double
        elif ch == "'":
            prev = text[i - 1] if i > 0 else ''
            result.append('\u2019' if prev.isalnum() or prev in ')]}' else '\u2018')
        else:
            result.append(ch)
    return ''.join(result)


def parse_question(raw) -> dict | None:
    """
    Parse a raw question cell string into structured MOSS question fields.

    Multiple Choice format:
        <stem>

        W) <option>
        X) <option>
        Y) <option>
        Z) <option>

        ANSWER: W) <answer text> [EH] <subcat>

    Short Answer format:
        <stem>

        ANSWER: <answer text> [EH] <subcat>
    """
    if not raw:
        return None
    text = str(raw).strip()

    # Split on blank lines
    blocks = re.split(r"\n[ \t]*\n", text)

    # ANSWER label may have typos ("ANSEWR:", "ASNWER:") — match any word + colon
    ANSWER_RE = re.compile(r"^[A-Za-z]+:\s*", re.MULTILINE)

    def strip_tags(s: str) -> str:
        """Remove trailing [TAG] editorial/subcat markers like [EH], [PB], [RC]."""
        return re.sub(r"\s*\[[A-Z&0-9]+\]\s*\S*\s*$", "", s).strip()

    # --- Multiple Choice ---
    if len(blocks) >= 3:
        stem = blocks[0].strip()
        options_block = blocks[1].strip()
        answer_block = blocks[2].strip()

        option_matches = re.findall(
            r"^([WXYZ])\)\s*(.+?)(?=\n[WXYZ]\)|$)",
            options_block,
            re.MULTILINE | re.DOTALL,
        )
        if len(option_matches) == 4:
            options = [m[1].strip() for m in option_matches]
            ans_match = ANSWER_RE.match(answer_block)
            if ans_match:
                letter_match = re.match(r"([WXYZ])\)", answer_block[ans_match.end() :])
                if letter_match:
                    return {
                        "question_style": "MULTIPLE_CHOICE",
                        "question_text": smartify_quotes(stem),
                        "options": [smartify_quotes(o) for o in options],
                        "correct_answer": letter_match.group(1).upper(),
                    }

    # --- Short Answer ---
    if len(blocks) >= 2:
        stem = blocks[0].strip()
        answer_block = blocks[-1].strip()
        ans_match = ANSWER_RE.match(answer_block)
        if ans_match:
            answer = strip_tags(answer_block[ans_match.end() :])
            return {
                "question_style": "SHORT_ANSWER",
                "question_text": smartify_quotes(stem),
                "options": [],
                "correct_answer": smartify_quotes(answer),
            }

    # Fallback — couldn't parse structure
    print(f"  WARNING: Could not parse question text, storing raw", file=sys.stderr)
    return {
        "question_style": "SHORT_ANSWER",
        "question_text": smartify_quotes(text),
        "options": [],
        "correct_answer": "",
    }


def get_fill_rgb(cell) -> str | None:
    try:
        fg = cell.fill.fgColor
        # If the color is stored as an RGB value, return it directly
        if fg.type == "rgb" and fg.rgb not in ("00000000", "FF000000"):
            return fg.rgb
        # For solid fills, Excel sometimes stores the color in bgColor
        bg = cell.fill.bgColor
        if bg.type == "rgb" and bg.rgb not in ("00000000", "FF000000"):
            return bg.rgb
        return None
    except Exception:
        return None


def read_compiling_sheet(wb) -> list[tuple[str, list[dict]]]:
    """
    Scan the Compiling sheet for blue-highlighted round blocks.

    The sheet is structured in 5-row blocks:
        i+0  Round label row  (col A = "ROUND ROBIN 1", blue fill = wanted)
        i+1  Question number row  ("1)", "1)", "2)", "2)", ...)
        i+2  Sheet name row  (B=Bio, C=Physics, ...)  ← INDIRECT first arg
        i+3  Question-type detection formulas (IFERROR/MC/SA)
        i+4  Question text formula row  =INDIRECT(B3&"!"&B$86&$CR5)

    The INDIRECT formula resolves as:
        sheet name  = row i+2, same column as formula cell
        col letter  = row 86 (fixed), same column as formula cell
        subject row = column CR of the formula row (same for all slots in a round)

    Returns list of (round_label, columns) where columns is a list of:
        {
            "question_type": "TOSSUP" | "BONUS",
            "category": "BIOLOGY" | ...,
            "sheet_name": "Bio" | ...,
            "excel_row": int,   # 1-indexed row in subject sheet
            "col_0idx": int,    # 0-indexed column in subject sheet
        }
    """
    sheet = wb["Compiling"]
    rows = list(sheet.iter_rows(values_only=False))
    wanted = []

    # Row 86 (0-indexed: 85) holds the subject-sheet column letter per question slot
    col_letter_row = rows[85] if len(rows) > 85 else []
    i = 0
    while i < len(rows):
        col_a = rows[i][0]
        if col_a.value and get_fill_rgb(col_a) == BLUE_FILL:
            if i + 4 < len(rows):
                round_label = str(col_a.value).strip()
                formula_row = rows[i + 4]
                sheet_name_row = rows[i + 2]  # INDIRECT first arg: subject sheet name per column
                columns = []

                # CR column contains =CEILING((row()+5)/5,1) — a formula we can't
                # read without data_only. Replicate it: formula row is at Excel row
                # i+5 (1-indexed), so subject row = (i + 10) // 5.
                excel_row = (i + 10) // 5

                for j in range(1, len(formula_row)):  # skip col A
                    form_cell = formula_row[j]
                    if form_cell.value is None:
                        break  # end of question slots for this round

                    sn_cell = sheet_name_row[j] if j < len(sheet_name_row) else None
                    cl_cell = col_letter_row[j] if j < len(col_letter_row) else None
                    sheet_name = str(sn_cell.value).strip() if (sn_cell and sn_cell.value) else None
                    col_letter = str(cl_cell.value).strip() if (cl_cell and cl_cell.value) else None

                    if not sheet_name or not col_letter:
                        continue

                    category = SHEET_TO_CATEGORY.get(sheet_name)
                    if not category:
                        print(
                            f"  WARNING: Unknown sheet {sheet_name!r}",
                            file=sys.stderr,
                        )
                        continue

                    q_type = col_to_question_type(col_letter)
                    col_0idx = column_index_from_string(col_letter) - 1

                    columns.append(
                        {
                            "question_type": q_type,
                            "category": category,
                            "sheet_name": sheet_name,
                            "excel_row": excel_row,
                            "col_0idx": col_0idx,
                        }
                    )

                if columns:
                    wanted.append((round_label, columns))
                    print(f"Found: {round_label!r}  ({len(columns)} question slots)")

            i += 5
        else:
            i += 1

    return wanted


def convert(excel_path: str, output_dir: str):
    print(f"Loading {excel_path} ...")

    # Without data_only → formula strings visible (used for Compiling sheet)
    wb_formulas = openpyxl.load_workbook(excel_path)
    # With data_only → cached cell values (used for subject sheets)
    wb_data = openpyxl.load_workbook(excel_path, data_only=True)

    rounds = read_compiling_sheet(wb_formulas)
    if not rounds:
        print("No blue-highlighted rounds found in Compiling sheet.", file=sys.stderr)
        return

    # Cache subject sheet rows so we don't re-read repeatedly
    sheet_cache: dict[str, list] = {}

    def get_cell_value(sheet_name: str, excel_row: int, col_0idx: int):
        if sheet_name not in sheet_cache:
            sheet = wb_data[sheet_name]
            sheet_cache[sheet_name] = list(sheet.iter_rows(values_only=True))
        rows = sheet_cache[sheet_name]
        row_idx = excel_row - 1  # 0-indexed
        if row_idx < len(rows) and col_0idx < len(rows[row_idx]):
            return rows[row_idx][col_0idx]
        return None

    os.makedirs(output_dir, exist_ok=True)

    for round_label, columns in rounds:
        questions = []
        q_id = 1
        pair_id = 1

        for col in columns:
            raw = get_cell_value(col["sheet_name"], col["excel_row"], col["col_0idx"])
            parsed = parse_question(raw)

            if parsed and col["category"] == "MATH":
                parsed["question_text"] = parsed["question_text"].replace("-", "\u2212")
                parsed["correct_answer"] = parsed["correct_answer"].replace("-", "\u2212")
                parsed["options"] = [o.replace("-", "\u2212") for o in parsed["options"]]

            if parsed:
                questions.append(
                    {
                        "id": q_id,
                        "pair_id": pair_id,
                        "question_type": col["question_type"],
                        "category": col["category"],
                        "source": "SSB_2026",
                        **parsed,
                    }
                )
                q_id += 1

            # Advance pair_id after each bonus slot (whether or not it had content)
            if col["question_type"] == "BONUS":
                pair_id += 1

        safe_name = re.sub(r"[^\w\s]", "", round_label).strip().lower()
        safe_name = re.sub(r"\s+", "_", safe_name)
        output_path = os.path.join(output_dir, f"{safe_name}.json")

        packet = {
            "packet": round_label,
            "year": 2026,
            "questions": questions,
        }

        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(packet, f, indent=4, ensure_ascii=False)

        print(f"  -> {output_path}  ({len(questions)} questions, {pair_id - 1} pairs)")

    print(f"\nDone. {len(rounds)} packet(s) written to {output_dir}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Convert SSB Question Writing Excel to MOSS JSON packets"
    )
    parser.add_argument("--input", default=EXCEL_PATH, help="Path to the Excel file")
    parser.add_argument("--output", default=OUTPUT_DIR, help="Output directory")
    args = parser.parse_args()
    convert(args.input, args.output)
